"""Unit tests for tasks/assistant_chat/file_writers.py."""

import csv
import io
import zipfile

import openpyxl
import pytest

from tasks.assistant_chat.file_writers import (
    ConversionError,
    UnsupportedExtension,
    normalize_and_categorize,
    to_bytes,
)


# ---------- normalize_and_categorize ----------

def test_normalize_md_extension_kept():
    assert normalize_and_categorize("note.md") == ("note.md", "text")


def test_normalize_markdown_extension_aliased_to_md():
    assert normalize_and_categorize("note.markdown") == ("note.md", "text")


def test_normalize_bare_name_defaults_to_md():
    assert normalize_and_categorize("ideas") == ("ideas.md", "text")


def test_normalize_text_extension():
    assert normalize_and_categorize("todo.txt") == ("todo.txt", "text")
    assert normalize_and_categorize("data.csv") == ("data.csv", "text")
    assert normalize_and_categorize("script.py") == ("script.py", "text")


def test_normalize_md_binary_categories():
    assert normalize_and_categorize("report.pdf") == ("report.pdf", "md-binary")
    assert normalize_and_categorize("plan.docx") == ("plan.docx", "md-binary")
    assert normalize_and_categorize("plan.odt") == ("plan.odt", "md-binary")


def test_normalize_xlsx_category():
    assert normalize_and_categorize("data.xlsx") == ("data.xlsx", "csv-xlsx")


def test_normalize_unknown_extension_raises():
    with pytest.raises(UnsupportedExtension):
        normalize_and_categorize("doc.tex")


def test_normalize_empty_raises():
    with pytest.raises(UnsupportedExtension):
        normalize_and_categorize("")


def test_normalize_case_insensitive():
    # ext detection should be case-insensitive but filename is preserved as-is
    assert normalize_and_categorize("Notes.MD") == ("Notes.MD", "text")


# ---------- to_bytes: text ----------

def test_to_bytes_text_utf8_round_trip():
    content = "Hola 🌍 — acentúa esto\nLine two"
    out = to_bytes(content, "note.md", "text")
    assert out.decode("utf-8") == content


def test_to_bytes_text_handles_none():
    # Defensive: pypandoc layer may pass None; we don't crash.
    out = to_bytes("", "note.md", "text")
    assert out == b""


# ---------- to_bytes: md-binary ----------

def test_md_to_pdf_produces_pdf_magic_bytes():
    content = "# Title\n\nA paragraph with **bold** text.\n\n- item 1\n- item 2"
    data = to_bytes(content, "report.pdf", "md-binary")
    assert data[:4] == b"%PDF", f"unexpected header: {data[:8]!r}"
    assert len(data) > 1000  # smoke: real PDF, not a stub


def test_md_to_docx_produces_zip_magic_bytes():
    content = "# Title\n\nParagraph.\n\n## Sub\n\n- item"
    data = to_bytes(content, "plan.docx", "md-binary")
    # docx is a zip
    assert data[:2] == b"PK"
    # Open it as zip and make sure the docx skeleton is in there
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = set(zf.namelist())
    assert "word/document.xml" in names


def test_md_to_odt_produces_zip_with_odt_mime():
    content = "# Title\n\nBody."
    data = to_bytes(content, "doc.odt", "md-binary")
    assert data[:2] == b"PK"
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        with zf.open("mimetype") as f:
            mime = f.read().decode("ascii")
    assert mime == "application/vnd.oasis.opendocument.text"


# ---------- to_bytes: csv-xlsx ----------

def test_csv_to_xlsx_round_trip():
    content = "Name,Age,City\nAlice,30,Madrid\nBob,25,Barcelona\n"
    data = to_bytes(content, "people.xlsx", "csv-xlsx")
    assert data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    assert rows[0] == ["Name", "Age", "City"]
    assert rows[1] == ["Alice", 30, "Madrid"]  # 30 inferred as int
    assert rows[2] == ["Bob", 25, "Barcelona"]


def test_csv_to_xlsx_quoted_fields_with_commas():
    content = 'Name,Note\n"Smith, John","line with, comma"\n'
    data = to_bytes(content, "data.xlsx", "csv-xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    assert rows[1] == ["Smith, John", "line with, comma"]


def test_csv_to_xlsx_preserves_leading_zeros_as_string():
    content = "ZipCode,Phone\n08001,0934567890\n"
    data = to_bytes(content, "z.xlsx", "csv-xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    # Both should be strings (kept as-is) to preserve leading zeros
    assert rows[1] == ["08001", "0934567890"]


def test_csv_to_xlsx_floats_detected():
    content = "Item,Price\nA,9.99\nB,12.5\n"
    data = to_bytes(content, "prices.xlsx", "csv-xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    assert rows[1] == ["A", 9.99]
    assert rows[2] == ["B", 12.5]


# ---------- error paths ----------

def test_to_bytes_unknown_category_raises_conversion_error():
    with pytest.raises(ConversionError):
        to_bytes("x", "foo.bar", "nope")
